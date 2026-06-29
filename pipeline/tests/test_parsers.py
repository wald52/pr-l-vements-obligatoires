"""Tests des parsers : taxes affectées (JSON), Eurostat (xlsx généré), V&M (helpers)."""

from pathlib import Path

import openpyxl

from po_pipeline import parse_eurostat, parse_taxes_affectees, parse_vm_tome1

FIXTURES = Path(__file__).parent / "fixtures"


def test_parse_taxes_affectees():
    recs = parse_taxes_affectees.parse(FIXTURES / "taxes_affectees.json", 2022)
    noms = {r.nom for r in recs}
    assert any("CVEC" in n for n in noms)
    cvec = next(r for r in recs if "CVEC" in r.nom)
    assert cvec.categorie == "taxe affectée"
    assert cvec.beneficiaire == "CROUS"
    assert cvec.montant_eur == 165 * 1_000_000
    assert cvec.base_legale and "éducation" in cvec.base_legale


def test_parse_taxes_affectees_misaligned(tmp_path):
    """Structure OpenDataSoft désalignée (plf2024) : intitulé en empty4, € bruts."""
    import json

    rows = [
        {  # ligne valide : taxe nommée en empty4, montant en euros
            "nature_juridique_du_beneficiaire": "Secteur social",
            "empty3": "Régime général de la Sécurité Sociale",
            "empty4": "Redevance due par les titulaires de titres d'exploitation de mines",
            "empty5": 3600000.0,
            "empty8": 3600000.0,
            "reference_juridique": "code minier",
        },
        {  # ligne à ignorer : empty4 non textuel / non fiscal
            "empty3": "Total",
            "empty4": 99.0,
            "empty8": 57000000000.0,
        },
    ]
    path = tmp_path / "ta_misaligned.json"
    path.write_text(json.dumps(rows), encoding="utf-8")

    recs = parse_taxes_affectees.parse(path, 2024)
    assert len(recs) == 1
    rec = recs[0]
    assert rec.nom.startswith("Redevance due par les titulaires")
    assert rec.categorie == "taxe affectée"
    assert rec.beneficiaire == "Régime général de la Sécurité Sociale"
    assert rec.montant_eur == 3_600_000  # euros bruts, pas de conversion M€
    assert rec.base_legale == "code minier"


def test_parse_eurostat_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ESA code", "Tax name", "Sector", "2023", "2024"])
    ws.append(["D5", "Impôt sur le revenu", "S1311", "85000", "90000"])
    ws.append(["D2", "TVA", "S1311", "190000", "200000"])
    ws.append([None, None, None, None, None])  # ligne vide ignorée
    path = tmp_path / "ntl.xlsx"
    wb.save(path)

    recs = parse_eurostat.parse(path, 2024)
    assert len(recs) == 2
    tva = next(r for r in recs if r.nom == "TVA")
    assert tva.esa_code == "D2"
    assert tva.secteur == "S1311"
    assert tva.montant_eur == 200000 * 1_000_000  # millions -> euros, année 2024


def test_parse_eurostat_ntl_country_sheet(tmp_path):
    """Format NTL officiel : un onglet FR avec en-tête STO/DETAILS et années."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("FR")
    ws.append(["TABLE_IDENTIFIER:", "T0999"])           # ligne de métadonnées
    ws.append(["STO ▼", "DETAILS ▼", "Tax name", None, 2023, 2024])  # en-tête
    ws.append(["D211", "_T", "Taxes de type TVA", None, 190000, 200000])  # sous-total -> ignoré
    ws.append(["D211", "C01", "TVA", None, 199000, 206332])             # taxe individuelle
    ws.append(["D91", "C01", "Mutations à titre gratuit", None, 18000, 20800])
    path = tmp_path / "ntl.xlsx"
    wb.save(path)

    recs = parse_eurostat.parse(path, 2024, geo="FR")
    noms = {r.nom for r in recs}
    assert "TVA" in noms and "Mutations à titre gratuit" in noms
    assert "Taxes de type TVA" not in noms  # le sous-total _T est écarté
    tva = next(r for r in recs if r.nom == "TVA")
    assert tva.esa_code == "D211"
    assert tva.montant_eur == 206332 * 1_000_000  # colonne 2024, M€ -> euros


def test_vm_amount_and_interpret():
    assert parse_vm_tome1._parse_amount("204 000") == 204000 * 1_000_000
    assert parse_vm_tome1._parse_amount("texte") is None
    nom, lineno, montant = parse_vm_tome1._interpret(["1101", "Taxe sur la valeur ajoutée", "204 000"])
    assert nom == "Taxe sur la valeur ajoutée"
    assert lineno == "1101"
    assert montant == 204000 * 1_000_000
